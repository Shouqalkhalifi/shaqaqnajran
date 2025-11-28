from django.shortcuts import render, redirect
from django.conf import settings
from django.contrib.auth import authenticate, login as auth_login
from urllib.parse import quote
import re
import requests
from openpyxl import load_workbook


def home(request):
    selected_unit_id = request.GET.get("unit_id")

    context = {
        "selected_unit_id": selected_unit_id,
        "checkin": request.GET.get("checkin", ""),
        "checkout": request.GET.get("checkout", ""),
        "adults": request.GET.get("adults", ""),
        "children": request.GET.get("children", ""),
    }

    return render(request, "home.html", context)


def contact(request):
    return render(request, "contact.html")


def admin_login(request):
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")

        user = authenticate(request, username=username, password=password)

        if user is not None and user.is_staff:
            auth_login(request, user)
            request.session["marketing_admin_authenticated"] = True
            return redirect("whatsapp_broadcast")

        return render(request, "admin/login.html", {"error": "بيانات الدخول غير صحيحة"})

    return render(request, "admin/login.html")


def whatsapp_broadcast(request):
    if not request.session.get("marketing_admin_authenticated"):
        return redirect("admin_login")

    contacts = request.session.get("broadcast_contacts", [])

    if request.method == "POST":
        # زر تفريغ القائمة
        if request.POST.get("clear_contacts"):
            request.session["broadcast_contacts"] = []
            return render(request, "admin/whatsapp_broadcast.html", {"contacts": []})

        # زر إرسال لجميع العملاء باستخدام القائمة المخزنة
        if request.POST.get("send_all"):
            contacts = request.session.get("broadcast_contacts", [])

            if contacts:
                access_token = settings.WHATSAPP_ACCESS_TOKEN
                phone_number_id = settings.WHATSAPP_PHONE_NUMBER_ID

                api_url = f"https://graph.facebook.com/v20.0/{phone_number_id}/messages"
                headers = {
                    "Authorization": f"Bearer {access_token}",
                    "Content-Type": "application/json",
                }

                for c in contacts:
                    payload = {
                        "messaging_product": "whatsapp",
                        "to": c["phone"],
                        "type": "text",
                        "text": {
                            "preview_url": False,
                            "body": c["text"],
                        },
                    }

                    try:
                        response = requests.post(api_url, json=payload, headers=headers, timeout=10)
                        print("WHATSAPP API STATUS:", response.status_code)
                        print("WHATSAPP API RESPONSE:", response.text)
                    except Exception as e:
                        print("WHATSAPP API ERROR:", e)
                        continue

            # بعد الإرسال نحافظ على القائمة للعرض فقط
            return render(request, "admin/whatsapp_broadcast.html", {"contacts": contacts})

        # رفع ملف جديد وتجهيز القائمة فقط بدون إرسال
        excel_file = request.FILES.get("excel_file")
        message = request.POST.get("message", "")

        if excel_file and message:
            new_contacts = []

            wb = load_workbook(excel_file, data_only=True)
            sheet = wb.active

            for row in sheet.iter_rows(min_row=2, values_only=True):
                name = row[0]
                phone = row[1]

                if not phone:
                    continue

                cleaned = re.sub(r"\D", "", str(phone))

                if cleaned.startswith("0"):
                    cleaned = "966" + cleaned[1:]

                if not cleaned.startswith("966"):
                    cleaned = "966" + cleaned

                personalized_text = str(message).replace("{name}", str(name) if name else "")
                encoded_text = quote(personalized_text)
                whatsapp_url = f"https://wa.me/{cleaned}?text={encoded_text}"

                new_contacts.append({
                    "name": name,
                    "phone": cleaned,
                    "whatsapp_url": whatsapp_url,
                    "text": personalized_text,
                })

            contacts = new_contacts
            request.session["broadcast_contacts"] = contacts

    return render(request, "admin/whatsapp_broadcast.html", {"contacts": contacts})
